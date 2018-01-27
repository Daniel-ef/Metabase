import os
import smtplib
from email import encoders
from email.mime.multipart import MIMEMultipart, MIMEBase
from email.mime.text import MIMEText
from urllib.parse import quote
from urllib.request import urlretrieve

import openpyxl
from openpyxl import load_workbook


def download_file(start_date, end_date, id, file_name):

    if file_name in os.listdir("."):
        print("File exists")
        return

    print("Downloading...")
    request_string = 'https://metabase.foxford.ru/public/question/782ef688-5f95-41ba-8cab-c285650cca86.xlsx?parameters='
    parametrs = '[{"type":"date/single","target":["variable",["template-tag","startDate"]],"value":"%s"},' \
              '{"type":"date/single","target":["variable",["template-tag","endDate"]],"value":"%s"},' \
              '{"type":"category","target":["variable",["template-tag","user_id"]],"value":"%s"},' \
              '{"type":"category","target":["variable",["template-tag","school_year_id"]],"value":"6"}' \
                     ']' % (start_date, end_date, id)

    url = request_string + quote(parametrs)
    urlretrieve(url, file_name)
    print("File downloaded")


def hide_columns(file_name):
    columns_to_delete = [1, 2, 4, 5]

    def hide_column(ws, column_id):
        if isinstance(column_id, int):
            column_id = openpyxl.utils.get_column_letter(column_id)
        column_dimension = ws.column_dimensions[column_id]
        column_dimension.hidden = True

    wb = load_workbook(file_name)
    ws = wb.active
    for col_num in columns_to_delete:
        hide_column(ws, col_num)

    wb.save(file_name)
    print("Columns hiden")


def send_email(filename, email):
    with open(filename, "rb") as fp:
        attachment = fp.read()
    with open("msg_body", encoding="utf-8") as fp:
        body = MIMEText(fp.read(), "plain")
    with open("auth", "r") as fp:
        login, password = fp.read().split()

    # Compose attachment
    part = MIMEBase('application', "octet-stream")
    part.set_payload(attachment)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % filename)

    # Compose message
    msg = MIMEMultipart()
    msg["Subject"] = "Фоксфорд. Еженедельный отчёт"
    msg["From"] = login
    msg["To"] = email

    msg.attach(part)
    msg.attach(body)

    # Establish connection
    server = smtplib.SMTP_SSL()
    server_addres = login.split("@")[1]
    server.connect('smtp.%s' % server_addres)
    server.login(login, password)

    # Send mail
    server.sendmail(msg["From"], msg["To"], msg.as_string())
    server.quit()
    print("Mail to {} sent".format(email))


if __name__ == "__main__":
    print("Put dates like: 2017-01-18")
    start_date, end_date = input("Start date: "), input("End date: ")

    with open("email-id", "r") as fp:
        e_id = list(map(lambda x: x.split(":"), fp.read().split()))

    for email, child_id in e_id:
        file_name = "{}---{}---{}.xlsx".format(start_date, end_date, child_id)

        download_file(start_date, end_date, child_id, file_name)

        hide_columns(file_name)

        send_email(file_name, email)


